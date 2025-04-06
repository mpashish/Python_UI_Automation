import inspect
import openpyxl
import pytest
import logging
from selenium.webdriver.support import expected_conditions as EC

from selenium.webdriver.support.wait import WebDriverWait



# @pytest.mark.usefixtures("setup")
class BaseClass:

    def verify_page_title(self, expected_title):
        assert self.driver.title in expected_title

    def wait_for_page_load(self, seconds):
        self.driver.implicitly_wait(seconds)

    def wait_until_element_presence(self, locator):
        element = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(locator))

    def verify_page_title_not_in(self, expected_title):
        assert self.driver.title not in expected_title

    def get_log_object(self):
        loggerName = inspect.stack()[1][3]  #This is alternate of below method to get current TC name

        # loger = logging.getLogger(__name__)

        loger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler("D:\\OneDrive - Fulcrum Digital\\Desktop\\Python\\Frameworks\\Python_UI_Automation\\logs\\appLogs.log")
        logFormatter  = logging.Formatter("%(asctime)s : %(levelname)s :%(name)s :%(message)s")
        fileHandler.setFormatter(logFormatter)
        loger.addHandler(fileHandler)
        loger.setLevel(logging.INFO)
        return loger

    def read_excel_as_list(self):
        excel_data_as_dictionary_list = []
        wb = openpyxl.load_workbook("D:\\OneDrive - Fulcrum Digital\\Desktop\\Python\\Frameworks\\Python_UI_Automation\\test_data\\PracticeExcel.xlsx")
        sheet = wb['Sheet1']
        rows = sheet.max_row
        columns = sheet.max_column
        for i in range(1, rows + 1):
            Dict = {}
            for j in range(1, columns + 1):
                columnName = sheet.cell(1, j).value
                if i == 1:
                    pass
                else:
                    Dict[columnName] = sheet.cell(row=i, column=j).value

            for k in excel_data_as_dictionary_list:
                if len(k) == 0:
                    excel_data_as_dictionary_list.remove(k)
            excel_data_as_dictionary_list.append(Dict)
        return excel_data_as_dictionary_list

    global excel_record
    def get_data_for(self, record_key, record_value):
        excel_record = {}
        data_list = self.read_excel_as_list()
        for item in data_list:
            if record_key in item.keys() and record_value in item.values():
                excel_record.update(item)
        return excel_record
